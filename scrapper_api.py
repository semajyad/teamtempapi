from __future__ import annotations
import os, re, time, json, uuid, tempfile, hashlib
from io import BytesIO
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import httpx
from bs4 import BeautifulSoup
from fastapi import FastAPI, HTTPException, Query, Path, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse

# ---------------------- CONFIG ----------------------
APP_VERSION = "file-8.0.0"
USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

# default list shown in your screenshot - guarantees UI is populated on fresh deploy
DEFAULT_SOURCES = [
    {"url": "https://teamtempapp.herokuapp.com/bvc/sDtXkQWe", "tribe": "SNZ Protect and Grow My Business"},
    {"url": "https://teamtempapp.herokuapp.com/bvc/qeGUw5jY", "tribe": "SNZ Service Excellence"},
    {"url": "https://teamtempapp.herokuapp.com/bvc/JvOZKWkI", "tribe": "SNZ Claims & Distribution"},
    {"url": "https://teamtempapp.herokuapp.com/bvc/HJJOxKb07".replace("x",""), "tribe": "SNZ Insurance Tech"},  # HJJOkb07
    {"url": "https://teamtempapp.herokuapp.com/bvc/Js7ly9PN", "tribe": "SNZ Run and Core Platforms"},
    {"url": "https://teamtempapp.herokuapp.com/bvc/mCO67NyY", "tribe": "SNZ Infrastructure"},
    {"url": "https://teamtempapp.herokuapp.com/bvc/DiMD4ZvL", "tribe": "SNZ Data CoE"},
]

# storage file inside the dyno/container
SOURCES_FILE = os.getenv("TEAMTEMP_SOURCES_FILE", "teamtemp_sources.json")
# optional seed or persistence via Heroku config var
SOURCES_JSON = os.getenv("SOURCES_JSON", "")

CACHE_TTL = int(os.getenv("CACHE_TTL_SEC", "600"))
HEROKU_APP_NAME = os.getenv("HEROKU_APP_NAME")
HEROKU_API_KEY  = os.getenv("HEROKU_API_KEY")

# -------------------- DATA MODEL --------------------
@dataclass
class Record:
    date: str
    team: str
    value: float
    tribe: str
    responses: Optional[int] = None
    min_value: Optional[float] = None
    max_value: Optional[float] = None

# ---------------------- APP ------------------------
app = FastAPI(title="TeamTemp Historical Scraper API", version=APP_VERSION)
app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

_client = httpx.Client(headers={"User-Agent": USER_AGENT}, timeout=30.0, follow_redirects=True)
_cache: Dict[str, object] = {"ts": 0.0, "data": []}

# ------------------ FILE PERSISTENCE ----------------
def _make_id(url: str) -> str:
    # deterministic id per URL so remove remains stable across restarts
    return hashlib.sha1(url.strip().encode("utf-8")).hexdigest()[:12]

def _atomic_write_json(path: str, data: object) -> None:
    d = os.path.dirname(path) or "."
    os.makedirs(d, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", delete=False, dir=d, encoding="utf-8") as tmp:
        json.dump(data, tmp, ensure_ascii=False, indent=2)
        tmp.flush(); os.fsync(tmp.fileno())
        tmp_path = tmp.name
    os.replace(tmp_path, path)

def _norm_row(r) -> Optional[dict]:
    if not isinstance(r, dict): return None
    url = str(r.get("url","")).strip()
    if not url: return None
    tribe = str(r.get("tribe","")).strip()
    return {
        "id": r.get("id") or _make_id(url),
        "url": url,
        "tribe": tribe,
        "created_ts": float(r.get("created_ts") or time.time()),
    }

def _ensure_rows(rows) -> List[dict]:
    # de-dupe by URL and keep stable id
    seen = set(); out: List[dict] = []
    for r in rows or []:
        rr = _norm_row(r)
        if not rr: continue
        if rr["url"] in seen: continue
        seen.add(rr["url"]); out.append(rr)
    out.sort(key=lambda x: (x["tribe"].lower(), x["url"].lower()))
    return out

def _mirror_to_heroku_config(rows) -> bool:
    if not (HEROKU_APP_NAME and HEROKU_API_KEY): return False
    try:
        httpx.patch(
            f"https://api.heroku.com/apps/{HEROKU_APP_NAME}/config-vars",
            headers={
                "Accept": "application/vnd.heroku+json; version=3",
                "Authorization": f"Bearer {HEROKU_API_KEY}",
                "Content-Type": "application/json",
            },
            json={"SOURCES_JSON": json.dumps(rows)},
            timeout=20.0,
        ).raise_for_status()
        return True
    except Exception:
        return False

def _initial_rows() -> List[dict]:
    # 1) If SOURCES_JSON is set, trust it completely
    if SOURCES_JSON:
        try:
            return _ensure_rows(json.loads(SOURCES_JSON))
        except Exception:
            pass
    # 2) If file exists, use it
    if os.path.exists(SOURCES_FILE):
        try:
            with open(SOURCES_FILE, "r", encoding="utf-8") as f:
                return _ensure_rows(json.load(f))
        except Exception:
            pass
    # 3) Seed from DEFAULT_SOURCES so UI is never empty on fresh deploy
    return _ensure_rows(DEFAULT_SOURCES)

def _read_sources_file() -> List[dict]:
    rows = _initial_rows()
    # ensure the file exists for later writes
    if not os.path.exists(SOURCES_FILE):
        _atomic_write_json(SOURCES_FILE, rows)
    return rows

def _write_sources_file(rows: List[dict]) -> List[dict]:
    rows = _ensure_rows(rows)
    _atomic_write_json(SOURCES_FILE, rows)
    _mirror_to_heroku_config(rows)  # optional persist across dyno restarts
    return rows

def list_sources() -> List[dict]:
    return _read_sources_file()

def add_source(url: str, tribe: str) -> dict:
    rows = _read_sources_file()
    # replace if same URL exists
    rows = [r for r in rows if r["url"] != url.strip()]
    row = {"id": _make_id(url), "url": url.strip(), "tribe": tribe.strip(), "created_ts": time.time()}
    rows.append(row)
    _write_sources_file(rows)
    return row

def delete_source(sid: str) -> bool:
    rows = _read_sources_file()
    new = [s for s in rows if s["id"] != sid]
    if len(new) == len(rows): return False
    _write_sources_file(new)
    return True

# ------------------- SCRAPING -----------------------
HISTORICAL_RE = re.compile(
    r"var\s+historical_data\s*=\s*new\s+google\.visualization\.DataTable\(\s*(\{.*?\})\s*(?:,\s*[^)]*)?\)\s*;",
    re.IGNORECASE | re.DOTALL,
)
DATE_STR_RE = re.compile(r"^\s*Date\((\d{4}),\s*(\d{1,2}),\s*(\d{1,2}).*?\)\s*$")
STATS_RE = re.compile(
    r"Min:\s*([0-9]+(?:\.[0-9]+)?)\s*,\s*Max:\s*([0-9]+(?:\.[0-9]+)?)\s*,\s*([0-9]+)\s+Responses?",
    re.IGNORECASE,
)

def _fetch_html(url: str) -> str:
    r = _client.get(url)
    r.raise_for_status()
    return r.text

def _extract_payload(html: str) -> Optional[dict]:
    m = HISTORICAL_RE.search(html)
    if not m:
        soup = BeautifulSoup(html, "html.parser")
        m = HISTORICAL_RE.search("\n".join(s.get_text("\n", strip=False) for s in soup.find_all("script")))
        if not m:
            return None
    obj = m.group(1)
    try:
        return json.loads(obj)
    except json.JSONDecodeError:
        try:
            return json.loads(obj.replace("'", '"'))
        except Exception:
            return None

def _date_from_cell(v) -> Optional[str]:
    if isinstance(v, str):
        mm = DATE_STR_RE.match(v)
        if mm:
            y, mo, d = int(mm.group(1)), int(mm.group(2)) + 1, int(mm.group(3))
            return f"{y:04d}-{mo:02d}-{d:02d}"
    return None

def _parse_stats(fmt: str) -> Tuple[Optional[int], Optional[float], Optional[float]]:
    if not fmt:
        return None, None, None
    m = STATS_RE.search(fmt)
    if not m:
        return None, None, None
    try:
        min_v = float(m.group(1)); max_v = float(m.group(2)); resp = int(m.group(3))
        return resp, min_v, max_v
    except Exception:
        return None, None, None

def _rows_to_records(payload: dict, tribe: str) -> List[Record]:
    cols = payload.get("cols", [])
    rows = payload.get("rows", [])
    if not cols or not rows: return []
    labels = [str(c.get("label") or c.get("id") or f"col{i}") for i, c in enumerate(cols[1:], start=1)]
    if labels and labels[-1].strip().lower() == "average":
        labels = labels[:-1]
    out: List[Record] = []
    for row in rows:
        cells = row.get("c", [])
        if not cells: continue
        date_iso = _date_from_cell(cells[0].get("v")) or time.strftime("%Y-%m-%d")
        for j, team in enumerate(labels, start=1):
            if j >= len(cells): continue
            cell = cells[j]
            if not isinstance(cell, dict): continue
            v = cell.get("v")
            if v is None: continue
            fmt = cell.get("f") or ""
            responses, min_v, max_v = _parse_stats(fmt)
            try:
                out.append(
                    Record(
                        date=date_iso,
                        team=team,
                        value=float(v),
                        tribe=tribe,
                        responses=responses,
                        min_value=min_v,
                        max_value=max_v,
                    )
                )
            except Exception:
                pass
    return out

def scrape_one(url: str, tribe: str) -> List[Record]:
    html = _fetch_html(url)
    payload = _extract_payload(html)
    return _rows_to_records(payload, tribe) if payload else []

# --------------------- FRONTEND ---------------------
INDEX_HTML = """
<!doctype html><html><head>
<meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
<meta http-equiv="Cache-Control" content="no-store, no-cache, must-revalidate"/>
<title>TeamTemp sources</title>
<style>
body{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;margin:24px}
.row{display:flex;gap:8px;margin-bottom:12px}
input{padding:8px} #url{flex:2} #tribe{flex:1}
button{padding:8px 12px;cursor:pointer}
table{border-collapse:collapse;width:100%;margin-top:12px}
th,td{border-bottom:1px solid #eee;text-align:left;padding:8px}
code{font-family:ui-monospace,Menlo,Consolas,monospace}
.badge{font-size:12px;color:#666}
</style></head><body>
<h1>TeamTemp sources</h1>
<div class="row">
  <input id="url" type="url" placeholder="https://teamtempapp.herokuapp.com/bvc/...">
  <input id="tribe" type="text" placeholder="Tribe">
  <button id="add">Add</button>
  <button id="refresh">Refresh data</button>
  <button id="dl">Download Excel</button>
</div>
<div id="msg" class="badge"></div>
<table><thead><tr><th>URL</th><th>Tribe</th><th>Actions</th></tr></thead><tbody id="rows"></tbody></table>
<script>
const el=(id)=>document.getElementById(id);
const api=(p,opt={})=>{
  const sep = p.includes("?") ? "&" : "?";
  return fetch(p+sep+"ts="+Date.now(), Object.assign({
    cache:"no-store",
    headers:{ "Content-Type":"application/json", "Cache-Control":"no-store" }
  },opt));
};
function render(list){
  list.sort((a,b)=> (a.tribe||"").localeCompare(b.tribe||"") || a.url.localeCompare(b.url));
  const tb = el('rows'); tb.innerHTML='';
  list.forEach(s=>{
    const tr=document.createElement('tr');
    tr.innerHTML = `<td><code>${s.url}</code></td><td>${s.tribe||''}</td>
      <td><button data-id="${s.id}" class="rm">Remove</button></td>`;
    tb.appendChild(tr);
  });
  document.querySelectorAll('.rm').forEach(b=>{
    b.onclick = async ()=>{
      await api('/sources/'+b.dataset.id,{method:'DELETE'});
      load();
    };
  });
}
async function load(retries=4, delay=250){
  try{
    const r = await api('/sources');
    if(!r.ok) throw new Error("bad status");
    const js = await r.json();
    render(js.sources||[]);
    el('msg').textContent = '';
  }catch(e){
    if(retries>0){
      el('msg').textContent = 'Loading… retrying';
      setTimeout(()=>load(retries-1, Math.min(2000, delay*2)), delay);
    }else{
      el('msg').textContent = 'Failed to load sources';
    }
  }
}
el('add').onclick = async ()=>{
  const url = el('url').value.trim(), tribe = el('tribe').value.trim();
  if(!url){ el('msg').textContent = 'Enter URL'; return; }
  const r = await api('/sources',{method:'POST',body:JSON.stringify({url,tribe})});
  if(!r.ok){ el('msg').textContent = 'Failed: '+await r.text(); return; }
  el('url').value=''; el('tribe').value='';
  load();
};
el('refresh').onclick = async ()=>{ el('msg').textContent='Refreshing…'; await api('/data?force=true'); el('msg').textContent='Refreshed'; };
el('dl').onclick = ()=>{ window.location.href='/export.xlsx?force=true'; };
load();
</script>
</body></html>
"""

# ----------------------- API -----------------------
def _no_store(resp: Response):
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"

@app.get("/", response_class=HTMLResponse)
def index() -> HTMLResponse:
    return HTMLResponse(INDEX_HTML)

@app.get("/version")
def version():
    return {"version": APP_VERSION, "storage": "file with optional heroku config var mirror"}

@app.get("/sources")
def sources_list(response: Response):
    _no_store(response)
    return {"sources": list_sources()}

@app.post("/sources")
def sources_add(payload: dict):
    url = str(payload.get("url","")).strip()
    tribe = str(payload.get("tribe","")).strip()
    if not url: raise HTTPException(400, "url required")
    row = add_source(url, tribe)
    _cache["data"] = []  # invalidate cache
    return row

@app.delete("/sources/{sid}")
def sources_delete(sid: str = Path(...)):
    if not delete_source(sid): raise HTTPException(404, "Not found")
    _cache["data"] = []
    return {"ok": True}

@app.get("/data")
def get_data(force: bool = Query(False), response: Response = None):
    if response: _no_store(response)
    now = time.time()
    if not force and (now - float(_cache.get("ts", 0))) < CACHE_TTL and isinstance(_cache.get("data"), list) and _cache["data"]:
        return _cache["data"]
    merged: List[Dict[str, object]] = []
    for s in list_sources():
        try:
            for rec in scrape_one(s["url"], s.get("tribe","")):
                merged.append(rec.__dict__)
        except Exception:
            pass
    _cache["ts"] = now
    _cache["data"] = merged
    return merged

def _excel_from_rows(rows: List[Dict[str, object]]) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "TeamTemp"
    header = ["tribe","team","date","value","responses","min","max"]
    ws.append(header)
    for r in rows:
        ws.append([
            r.get("tribe",""),
            r.get("team",""),
            r.get("date",""),
            r.get("value",""),
            r.get("responses",""),
            r.get("min_value",""),
            r.get("max_value",""),
        ])
    for i in range(1, len(header)+1):
        maxlen = max(len(str(ws.cell(row=r, column=i).value or "")) for r in range(1, ws.max_row+1))
        ws.column_dimensions[get_column_letter(i)].width = min(maxlen+2, 60)
    bio = BytesIO(); wb.save(bio); bio.seek(0); return bio

@app.get("/export.xlsx")
def export_excel(force: bool = Query(False), response: Response = None):
    if response: _no_store(response)
    if force or not _cache.get("data"):
        get_data(force=True)
    rows = _cache["data"] if isinstance(_cache["data"], list) else []
    stream = _excel_from_rows(rows)
    fname = f"teamtemp_{time.strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        stream,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("teamtemp_multi_sources_with_tribe:app", host="0.0.0.0", port=int(os.getenv("PORT", "8000")), reload=False)
